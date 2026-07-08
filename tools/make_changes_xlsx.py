"""Generate the TWS Portfolio Changes workbook (one tab per portfolio)
from the dashboard's changes.js. Standing deliverable for Brian's team.
Usage: python tools/make_changes_xlsx.py [output.xlsx]
"""
import json, re, subprocess, sys, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
    r"C:\Users\decke\OneDrive\Desktop\Skool\Main TWS Docs 2026") / f"TWS_Portfolio_Changes_{datetime.date.today():%Y-%m-%d}.xlsx"

# Parse changes.js + portfolio names via node (concat into one script so consts share scope)
import tempfile
extract = (
    "\nconst out={};for(const pid of Object.keys(CHANGES)){"
    "out[pid]={name:PORTFOLIOS[pid].name,layers:PORTFOLIOS[pid].layers.map(l=>l.name),rows:CHANGES[pid]}};"
    "console.log(JSON.stringify(out));\n"
)
src = (ROOT / "data.js").read_text(encoding="utf-8") + "\n" + (ROOT / "changes.js").read_text(encoding="utf-8") + extract
with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False, encoding="utf-8") as f:
    f.write(src); tmp = f.name
data = json.loads(subprocess.run(["node", tmp], capture_output=True, text=True, check=True).stdout)

FONT = "Arial"
INK, CREAM, TEAL, GOLD = "0D0F10", "F0ECE4", "57C8D8", "F0B429"
STATUS_FILL = {"kept": "EDEDED", "swap": "FDEBC8", "add": "D9F2F2"}
STATUS_LABEL = {"kept": "KEPT", "swap": "SWAPPED", "add": "NEW"}
thin = Border(bottom=Side(style="thin", color="D9D9D9"))

wb = Workbook()
wb.remove(wb.active)
order = ["def", "agg", "hc"]

for pid in order:
    p = data[pid]
    ws = wb.create_sheet(p["name"].replace(" Portfolio", ""))
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"TWS {p['name']} — What Changed (July 2026 Update)"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
    ws["A2"] = f"Generated {datetime.date.today():%B %d, %Y}. Previous = April 2026 site. Tickers are sector examples, not guarantees — see MidasX.ai. Educational purposes only."
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="666666")

    headers = ["Layer", "Prev Ticker", "Prev Option", "Prev Alloc %", "Change",
               "New Ticker", "New Option", "New Alloc %", "Alloc Δ (pts)"]
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(name=FONT, size=10, bold=True, color=CREAM)
        cell.fill = PatternFill("solid", start_color=INK)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")

    r = hr + 1
    for li, lname in enumerate(p["layers"]):
        rows = [x for x in p["rows"] if x["l"] == li]
        if not rows:
            continue
        lc = ws.cell(row=r, column=1, value=lname.upper())
        lc.font = Font(name=FONT, size=10, bold=True, color=INK)
        lc.fill = PatternFill("solid", start_color="DCE9EA")
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color="DCE9EA")
        layer_start = r + 1
        r += 1
        for x in rows:
            old, nw = x.get("old"), x.get("nw")
            vals = ["",
                    old["t"] if old else "—", old["opt"] if old else "—", old["w"] if old else None,
                    STATUS_LABEL[x["status"]],
                    nw["t"] if nw else "—", nw["opt"] if nw else "—", nw["w"] if nw else None,
                    f"=H{r}-D{r}" if (old and nw) else None]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(name=FONT, size=10,
                                 bold=(c in (2, 6)),
                                 strike=(c == 2 and x["status"] == "swap"),
                                 color=INK)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
                if c == 5:
                    cell.fill = PatternFill("solid", start_color=STATUS_FILL[x["status"]])
                    cell.font = Font(name=FONT, size=9, bold=True, color=INK)
                if c in (4, 8):
                    cell.number_format = '0"%"'
                if c == 9:
                    cell.number_format = '+0;-0;"—"'
            r += 1
        # layer subtotal (new allocation)
        st = ws.cell(row=r, column=1, value=f"{lname} total (new)")
        st.font = Font(name=FONT, size=9, bold=True, color="666666")
        tot = ws.cell(row=r, column=8, value=f"=SUM(H{layer_start}:H{r-1})")
        tot.font = Font(name=FONT, size=10, bold=True, color=INK)
        tot.number_format = '0"%"'
        tot.alignment = Alignment(horizontal="center")
        r += 2

    gt = ws.cell(row=r, column=1, value="PORTFOLIO TOTAL (new)")
    gt.font = Font(name=FONT, size=11, bold=True, color=INK)
    tot_cells = []
    rr = hr + 1
    for li, lname in enumerate(p["layers"]):
        # find each subtotal row by scanning column A
        pass
    # simpler: grand total = sum of all New Alloc entries
    gtot = ws.cell(row=r, column=8, value=f"=SUMPRODUCT((H{hr+1}:H{r-1})*(A{hr+1}:A{r-1}=\"\"))")
    gtot.number_format = '0"%"'
    gtot.font = Font(name=FONT, size=11, bold=True, color=INK)
    gtot.alignment = Alignment(horizontal="center")

    widths = [26, 12, 12, 12, 11, 12, 12, 12, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{hr+1}"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"saved: {OUT}")
